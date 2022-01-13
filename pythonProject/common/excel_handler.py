import openpyxl


class ExcelHandler():

    def __init__(self, file):
        self.file = file

    def open_sheet(self, name):
        wb = openpyxl.load_workbook(self.file)
        sheet = wb[name]
        return sheet

    def header(self, name):
        sheet = self.open_sheet(name)
        headers = []
        for i in sheet[1]:
            headers.append(i.value)
        return headers

    def read(self, name):
        sheet = self.open_sheet(name)
        rows = list(sheet.rows)

        data = []
        for row in rows[1:]:
            row_data = []
            for cell in row:
                row_data.append(cell.value)
                data_dict = dict(zip(self.header(name), row_data))
            data.append(data_dict)
        return data

    def write(self, name, row, column, data):
        wb = openpyxl.load_workbook(self.file)
        sheet = wb[name]
        sheet.cell(row, column).value = data
        wb.save(self.file)
        wb.close()


if __name__ == '__main__':
    excel = ExcelHandler(r'e:\cases.xlsx')
    sheet = excel.read("Sheet1")
    print(sheet)

